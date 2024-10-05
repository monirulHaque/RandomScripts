import sys
import subprocess

# implement pip as a subprocess:
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 
'openpyxl'])
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 
'selenium'])
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 
'maskpass'])
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 
'webdriver-manager'])

import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import maskpass  # to hide the password
from os import path


# Initializing chrome driver in selenium bot
# chrome_options = webdriver.ChromeOptions()
# driver = webdriver.Chrome()
driver = webdriver.Chrome(ChromeDriverManager().install())

#Navigating to Advising Area
driver.get("https://vjudge.net/contest/632171#rank")
try:
    login_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "a.nav-link.login"))
    )
except:
    print("ERROR: LOGIN BUTTON NOT FOUND!")
login_button.click()
time.sleep(1)
username_input = driver.find_element(By.ID, "login-username")
password_input = driver.find_element(By.ID, "login-password")
username_input.send_keys("monirultanvir63@gmail.com")
pwd = maskpass.advpass()
password_input.send_keys(pwd)
login_button = driver.find_element(By.ID, "btn-login")
login_button.click()

# Waiting for the page to load 
try:
    table_rows = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='contest-rank-table']/tbody/tr"))
    )
except:
    print("ERROR: CONTEST TABLE NOT FOUND!")

# Creating the sheet
filename = "contest.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "UserName"
ws['B1'] = "Profile URL"
ws['C1'] = "Solved"
ws['D1'] = "Rank"
ws['E1'] = "Considering Upsolve"
ws['F1'] = "Rank Afterwards"

time.sleep(2)
# Getting the values
table_rows = driver.find_elements(By.XPATH, "//*[@id='contest-rank-table']/tbody/tr")
row_no = 2
for row in table_rows:
    profile_url = row.find_element(By.CSS_SELECTOR, "a")
    score = row.find_element(By.CLASS_NAME, "solved").text
    ws[f'A{row_no}'] = profile_url.text
    ws[f'B{row_no}'] = profile_url.get_attribute('href')
    ws[f'C{row_no}'] = score
    ws[f'D{row_no}'] = row.find_element(By.CLASS_NAME, "rank").text
    row_no += 1

wb.save(filename)


wb = openpyxl.load_workbook(filename)
ws = wb.active
# Upsolve Count
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "btn-setting"))).click()
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//label[@class='btn btn-secondary btn-sm'][normalize-space()='Show'])[1]"))).click()
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//span[@aria-hidden='true'])[1]"))).click()
try:
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, "show-all-teams"))).click()
except:
    print('No button found!')
table_rows = driver.find_elements(By.XPATH, "//*[@id='contest-rank-table']/tbody/tr")
for row in table_rows:
    profile_url = row.find_element(By.CSS_SELECTOR, "a")
    # searching the person in the excel sheet
    found_row = -1
    for i in range(2, row_no):
        if ws[f'B{i}'].value == profile_url.get_attribute('href'):
            found_row = i
            break
    if found_row == -1:
        ws[f'A{row_no}'] = profile_url.text
        ws[f'B{row_no}'] = profile_url.get_attribute('href')
        ws[f'C{row_no}'] = 0
        ws[f'D{row_no}'] = ""
        ws[f'E{row_no}'] = row.find_element(By.CLASS_NAME, "solved").text
        ws[f'F{row_no}'] = row.find_element(By.CLASS_NAME, "rank").text
        row_no += 1
    else:
        ws[f'E{found_row}'] = row.find_element(By.CLASS_NAME, "solved").text
        ws[f'F{found_row}'] = row.find_element(By.CLASS_NAME, "rank").text
    
wb.save(filename)