# -*- coding: utf-8 -*-
"""
Created on Sun Dec  4 09:28:07 2022

@author: ext.monirul.haque
"""

import pandas as pd
from openpyxl import Workbook
import math
filename = 'section9'
df = pd.read_csv(f'{filename}.csv')
dic = df.to_dict()

wb = Workbook()
ws = wb.active
ws.append(['Group No', 'Topic No.','Member ID', 'Member Name'])

cellCounter = 2
for i in range(len(df)):
    mergeStart = cellCounter
    for key in dic:
        if key == 'Group No.':
            group = int(dic[key][i])
        elif key == 'Topic':
            topic = int(dic[key][i])
        elif type(dic[key][i]) is str:
            ws[f'D{cellCounter}'] = dic[key][i]
            cellCounter += 1
        elif not math.isnan(dic[key][i]):
            ws[f'C{cellCounter}'] = dic[key][i]
    ws[f'A{mergeStart}'] = group
    ws[f'B{mergeStart}'] = topic
    ws.merge_cells(f'A{mergeStart}:A{cellCounter-1}')
    ws.merge_cells(f'B{mergeStart}:B{cellCounter-1}')
wb.save(f'{filename}.xlsx')
