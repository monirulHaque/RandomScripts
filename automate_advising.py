import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import time
import maskpass  # to hide the password
from os import path

# Initializing chrome driver in selenium bot
chrome_options = webdriver.ChromeOptions()
driver = webdriver.Chrome()

#Navigating to Advising Area
driver.get("https://usis.bracu.ac.bd/academia/")
time.sleep(4)
username_input = driver.find_element(By.ID, "username")
password_input = driver.find_element(By.ID, "password")
username_input.send_keys("ext.monirul.haque@bracu.ac.bd")
pwd = maskpass.advpass()
password_input.send_keys(pwd)
login_button = driver.find_element(By.ID, "ctl00_leftColumn_ctl00_btnLogin")
login_button.click()
time.sleep(2)
driver.get("https://usis.bracu.ac.bd/academia/dashBoard/show#/academia/studentCourse/showAdvisingWithFilter")
time.sleep(2)
advising_year_selector = Select(driver.find_element(By.ID, "academiaYear"))
advising_year_selector.select_by_value("279288")
time.sleep(2)
advising_session_selector = Select(driver.find_element(By.ID, "academiaSession"))
advising_session_selector.select_by_value("627120")

#Advising Sheet to get the IDs
wb = openpyxl.load_workbook('advising_Spring2023.xlsx') 
ws = wb.active
time.sleep(2)
# while True:
    

for i in range(6,ws.max_row+1):
    
    # id_input = driver.find_element(By.ID, "idNo")
    id_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "idNo"))
    )
    id_input.clear()
    time.sleep(1)
    id_input.send_keys(ws[f'B{i}'].value)
    show_button = driver.find_element(By.ID, "show-details-button")
    show_button.click()
    time.sleep(5)
    #waitforinput
    print(ws[f'B{i}'].value)
    print('Remarks: ',end="")
    remarks = input()
    if remarks == 'close':
        break
    ws[f'H{i}'] = remarks
    time.sleep(1)

wb.save('advising_Spring2023.xlsx')
